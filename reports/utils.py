import io
from datetime import date, datetime, timezone
import openpyxl
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, Image, BaseDocTemplate, Frame, PageTemplate
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors, enums
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.charts.legends import Legend
from reportlab.lib.units import inch


# === TEMAS Y ESTILOS PARA REPORTES ===
class ReportTheme:
    PRIMARY = colors.HexColor('#004B87')    # Azul Institucional
    SECONDARY = colors.HexColor('#0072CE')  # Azul Brillante
    ACCENT = colors.HexColor('#63666A')     # Gris Profesional
    TEXT = colors.HexColor('#212529')       # Texto casi negro
    LIGHT_GREY = colors.HexColor('#F8F9FA') # Fondo de tablas/zebra
    BORDER = colors.HexColor('#DEE2E6')     # Bordes sutiles
    WHITE = colors.whitesmoke

    @classmethod
    def get_styles(cls):
        styles = getSampleStyleSheet()
        
        # Título Principal
        styles.add(ParagraphStyle(
            name='ModernTitle',
            parent=styles['Title'],
            fontName='Helvetica-Bold',
            fontSize=22,
            textColor=cls.PRIMARY,
            spaceAfter=20,
            alignment=enums.TA_LEFT
        ))
        
        # Encabezados de Sección
        styles.add(ParagraphStyle(
            name='ModernHeading',
            parent=styles['Heading2'],
            fontName='Helvetica-Bold',
            fontSize=16,
            textColor=cls.PRIMARY,
            spaceBefore=15,
            spaceAfter=10,
            borderPadding=(0, 0, 5, 0),
            borderColor=cls.SECONDARY,
            borderWidth=0, # Solo queremos el espacio
            alignment=enums.TA_LEFT
        ))

        # Subtítulos
        styles.add(ParagraphStyle(
            name='ModernSubHeading',
            parent=styles['Heading3'],
            fontName='Helvetica-Bold',
            fontSize=12,
            textColor=cls.SECONDARY,
            spaceBefore=10,
            spaceAfter=6
        ))

        # Texto Normal
        styles.add(ParagraphStyle(
            name='ModernNormal',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=10,
            textColor=cls.TEXT,
            leading=14,
            alignment=enums.TA_JUSTIFY
        ))

        # Estilo para celdas de tabla
        styles.add(ParagraphStyle(
            name='TableCell',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=9,
            textColor=cls.TEXT,
            alignment=enums.TA_CENTER
        ))

        return styles


def _create_styled_table(data, col_widths=None):
    """Crea una tabla con diseño moderno y filas alternadas."""
    t = Table(data, colWidths=col_widths)
    
    # Estilo base
    style = [
        ('BACKGROUND', (0, 0), (-1, 0), ReportTheme.PRIMARY),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('TOPPADDING', (0, 0), (-1, 0), 10),
        
        # Bordes y cuerpo
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, ReportTheme.BORDER),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, ReportTheme.LIGHT_GREY]),
    ]
    
    t.setStyle(TableStyle(style))
    return t


def header_footer(canvas, doc):
    """Dibuja el encabezado y pie de página en cada página."""
    canvas.saveState()
    
    # --- ENCABEZADO ---
    # Logo (si existe)
    logo_path = 'index/static/Template/ssa_logo.webp'
    try:
        canvas.drawImage(logo_path, 40, letter[1] - 70, width=120, height=45, preserveAspectRatio=True, mask='auto')
    except:
        pass
    
    # Título del reporte en el encabezado
    canvas.setFont('Helvetica-Bold', 10)
    canvas.setStrokeColor(ReportTheme.PRIMARY)
    canvas.setLineWidth(0.5)
    canvas.line(40, letter[1] - 75, letter[0] - 40, letter[1] - 75)
    
    canvas.setFillColor(ReportTheme.ACCENT)
    canvas.drawRightString(letter[0] - 40, letter[1] - 65, "SIFOR - Sistema de Formularios")
    
    # --- PIE DE PÁGINA ---
    canvas.line(40, 50, letter[0] - 40, 50)
    canvas.setFont('Helvetica', 8)
    fecha = datetime.now().strftime('%d/%m/%Y %H:%M')
    canvas.drawString(40, 40, f"Generado el: {fecha}")
    canvas.drawRightString(letter[0] - 40, 40, f"Página {doc.page}")
    
    canvas.restoreState()


def export_queryset_to_excel(queryset, filename='reporte', excluded_fields=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte"

    model = queryset.model
    opts = model._meta
    fields = [
        f for f in opts.concrete_fields
        if f.name not in (excluded_fields or [])
    ]

    # === TÍTULO ===
    title = opts.verbose_name_plural.title()
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(fields))
    cell = ws.cell(row=1, column=1, value=title)
    cell.font = Font(size=14, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")

    # === ENCABEZADOS ===
    headers = [f.verbose_name.title() for f in fields]
    ws.append([])  # fila vacía después del título
    ws.append(headers)

    # === FILAS ===
    for obj in queryset:
        row = []
        for field in fields:
            # Si el campo tiene choices → usar display
            if field.choices:
                value = getattr(obj, f"get_{field.name}_display")()
            else:
                value = getattr(obj, field.name)

            # Manejo de tipos
            if value is None:
                row.append('')
            elif isinstance(value, str):
                row.append(value)
            elif isinstance(value, (int, float, bool)):
                row.append(value)
            elif isinstance(value, (date, datetime)):
                fmt = "%d-%m-%Y %H:%M" if isinstance(value, datetime) else "%d-%m-%Y"
                row.append(value.strftime(fmt))
            else:
                row.append(str(value))

        ws.append(row)

    # === AUTOAJUSTE DE ANCHO DE COLUMNAS ===
    for i, col in enumerate(ws.columns, 1):
        max_length = 0
        col_letter = get_column_letter(i)  # más seguro que usar col[0]
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    # === RESPUESTA ===
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={filename}.xlsx'
    wb.save(response)
    return response


def export_generic_history_to_excel(queryset, filename='historial', sheet_title='Historial'):
    if not queryset.exists():
        raise ValueError("El queryset está vacío.")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title

    # Obtener campos base del modelo
    model = queryset.model
    base_fields = [f.name for f in model._meta.concrete_fields]

    # Tomar un ejemplo de data para inferir campos JSON (si existe)
    example_entry = queryset.first()
    json_data_keys = list(example_entry.data.keys()) if hasattr(example_entry, 'data') else []

    # Combinar encabezados
    headers = base_fields + json_data_keys
    ws.append(headers)

    # Escribir filas
    for entry in queryset:
        base_values = []
        for field in base_fields:
            value = getattr(entry, field)
            base_values.append(str(value) if value is not None else '')

        data_values = []
        for key in json_data_keys:
            value = entry.data.get(key, '')
            data_values.append(str(value))

        ws.append(base_values + data_values)

    # Preparar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={filename}.xlsx'
    wb.save(response)
    return response


def export_extended_queryset_to_excel(queryset, fields, filename='reporte', sheet_title='Datos'):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title

    # Escribir encabezados
    ws.append([field[0] for field in fields])  # Ejemplo: ("Código", "code")

    # Escribir filas
    for obj in queryset:
        row = []
        for label, key in fields:
            value = None

            if '__' in key:
                # Soporta relaciones como support_info__problem
                parts = key.split('__')
                current = obj
                for part in parts:
                    current = getattr(current, part, None)
                    if current is None:
                        break
                value = str(current) if current is not None else ''
            else:
                # Campos directos de Transaction u otros objetos
                if hasattr(obj, key):
                    value = getattr(obj, key)
                elif hasattr(obj, 'support_info') and key.startswith('support_info_'):
                    sub_key = key.replace('support_info_', '')
                    value = getattr(obj.support_info, sub_key, '')
                elif hasattr(obj, 'output_info') and key.startswith('output_info_'):
                    sub_key = key.replace('output_info_', '')
                    value = getattr(obj.output_info, sub_key, '')
                else:
                    value = ''

            row.append(str(value) if value is not None else '')

        ws.append(row)

    # Preparar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={filename}.xlsx'
    wb.save(response)
    return response


def export_queryset_to_excel_advance(queryset, filename='reporte', excluded_fields=None):
    """
    Exporta un queryset grande a Excel (.xlsx) de forma eficiente usando openpyxl en modo write_only.
    - Evita sobrecargar la RAM.
    - Normaliza datetimes timezone-aware a timezone-naive.
    """
    model = queryset.model
    opts = model._meta
    excluded_fields = excluded_fields or []

    # === Crear workbook optimizado ===
    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title="Reporte")

    # === Título ===
    title = opts.verbose_name_plural.title()
    ws.append([title])
    ws.append([])  # fila vacía

    # === Encabezados ===
    fields = [f for f in opts.concrete_fields if f.name not in excluded_fields]
    headers = [f.verbose_name.title() for f in fields]
    ws.append(headers)

    # === Escribir filas (sin cargar all en memoria) ===
    for obj in queryset.iterator(chunk_size=10000):
        row = []
        for field in fields:
            # Si el campo tiene choices → usar display
            if field.choices:
                value = getattr(obj, f"get_{field.name}_display")()
            else:
                value = getattr(obj, field.name)

            # Normalizar valores
            if value is None:
                row.append('')
            elif isinstance(value, (str, int, float, bool)):
                row.append(value)
            elif isinstance(value, datetime):
                # Convertir timezone-aware a naive de forma segura
                if value.tzinfo is not None:
                    try:
                        value = value.astimezone(timezone.utc).replace(tzinfo=None)
                    except Exception:
                        # fallback seguro: ignorar tzinfo
                        value = value.replace(tzinfo=None)
                row.append(value.strftime("%d-%m-%Y %H:%M"))
            elif isinstance(value, date):
                row.append(value.strftime("%d-%m-%Y"))
            else:
                row.append(str(value))

        ws.append(row)

    # === Guardar en memoria y enviar ===
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = StreamingHttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={filename}.xlsx'
    return response


import csv
from datetime import date, datetime
from django.http import StreamingHttpResponse


class Echo:
    """Helper class para streaming del CSV"""

    def write(self, value):
        return value


def export_queryset_to_csv_fast(queryset, filename='reporte', fields=None, delimiter=','):
    """
    Exporta cualquier queryset a CSV de forma ultrarrápida usando values().
    Soporta caracteres especiales (UTF-8 con BOM para Excel).
    """
    if fields is None:
        # Si no se especifican campos, tomar todos los del modelo
        model = queryset.model
        fields = [f.name for f in model._meta.concrete_fields]

    # Obtener datos planos
    qs = queryset.values(*fields)

    def normalize_value(value):
        if value is None:
            return ''
        elif isinstance(value, datetime):
            return value.strftime("%Y-%m-%d %H:%M:%S")
        elif isinstance(value, date):
            return value.strftime("%Y-%m-%d")
        return str(value)

    def csv_generator():
        pseudo_buffer = Echo()
        writer = csv.writer(pseudo_buffer, delimiter=delimiter)
        # Agregar BOM UTF-8 (para Excel)
        yield '\ufeff'
        yield writer.writerow(fields)
        for row in qs.iterator(chunk_size=20000):
            yield writer.writerow([normalize_value(v) for v in row.values()])

    response = StreamingHttpResponse(
        csv_generator(),
        content_type='text/csv; charset=utf-8'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
    return response


import openpyxl

from django.http import HttpResponse
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def export_form_responses_to_excel(form, filename=None):

    filename = filename or f"{form.title}_respuestas"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Respuestas"

    questions = form.questions.all()
    choices_by_question = {
        question.id: {str(choice.id): choice.choice for choice in question.choices.all()}
        for question in questions
    }

    def translate_answer_value(question_id, raw_value):
        choices_map = choices_by_question.get(question_id, {})
        if not choices_map or raw_value is None:
            return raw_value

        normalized_value = str(raw_value).strip()

        if "," in normalized_value:
            translated_parts = []
            for part in normalized_value.split(","):
                part = part.strip()
                translated_parts.append(choices_map.get(part, part))
            return ", ".join(translated_parts)

        return choices_map.get(normalized_value, raw_value)

    # ===== HEADERS FIJOS =====
    headers = [
        "Código Respuesta",
        "Correo",
        "Establecimiento",
        # "IP"
    ]

    # ===== AGREGAR PREGUNTAS COMO COLUMNAS =====
    for question in questions:
        headers.append(question.question)

    ws.append(headers)

    # ===== ESTILOS HEADER =====
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # ===== RESPUESTAS =====
    responses = (
        form.response_to
        .select_related('responder', 'responder__establecimiento')
        .prefetch_related('response__answer_to')
        .all()
    )

    for response in responses:

        responder = response.responder

        row = [
            response.response_code,
            # responder.username if responder else '',
            response.responder_email,
            responder.establecimiento.alias if responder and responder.establecimiento else '',
            # response.responder_ip
        ]

        # Diccionario pregunta_id -> respuesta
        answers_map = {}

        for answer in response.response.all():

            question_id = answer.answer_to.id

            if question_id not in answers_map:
                answers_map[question_id] = []

            answers_map[question_id].append(
                str(translate_answer_value(question_id, answer.answer))
            )

        # Agregar respuestas según orden de preguntas
        for question in questions:

            answers = answers_map.get(question.id, [])

            row.append(", ".join(answers))

        ws.append(row)

    # ===== AUTO WIDTH =====
    for column in ws.columns:

        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        ws.column_dimensions[column_letter].width = min(max_length + 5, 60)

    # ===== RESPONSE =====
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    response['Content-Disposition'] = (
        f'attachment; filename="{filename}.xlsx"'
    )

    wb.save(response)

    return response



def _create_pie_chart_drawing(data_dict):
    """Crea un Drawing con un gráfico de torta y su leyenda con colores institucionales."""
    if not data_dict:
        return None
    
    drawing = Drawing(400, 200)
    pc = Pie()
    pc.x = 60
    pc.y = 50
    pc.width = 120
    pc.height = 120
    pc.data = list(data_dict.values())
    pc.labels = [f"{v} ({(v/sum(pc.data)*100):.1f}%)" if sum(pc.data) > 0 else str(v) for v in data_dict.values()]
    
    color_palette = [
        ReportTheme.PRIMARY, ReportTheme.SECONDARY, colors.HexColor('#28A745'), 
        ReportTheme.ACCENT, colors.HexColor('#FFC107'), colors.HexColor('#17A2B8'),
        colors.HexColor('#DC3545'), colors.HexColor('#6610F2'), colors.HexColor('#E83E8C')
    ]
    for i in range(len(pc.data)):
        pc.slices[i].fillColor = color_palette[i % len(color_palette)]
        pc.slices[i].strokeColor = colors.white
        pc.slices[i].strokeWidth = 0.5

    legend = Legend()
    legend.x = 220
    legend.y = 170
    legend.dx = 8
    legend.dy = 8
    legend.fontName = 'Helvetica'
    legend.fontSize = 8
    legend.alignment = 'right'
    legend.columnMaximum = 12
    legend.colorNamePairs = [(pc.slices[i].fillColor, str(k)) for i, k in enumerate(data_dict.keys())]
    
    drawing.add(pc)
    drawing.add(legend)
    return drawing


def _create_bar_chart_drawing(data_dict):
    """Crea un Drawing con un gráfico de barras verticales moderno."""
    if not data_dict:
        return None
        
    drawing = Drawing(400, 220)
    bc = VerticalBarChart()
    bc.x = 50
    bc.y = 70
    bc.height = 125
    bc.width = 320
    bc.data = [list(data_dict.values())]
    bc.strokeColor = ReportTheme.BORDER
    bc.fillColor = ReportTheme.LIGHT_GREY
    
    bc.bars[0].fillColor = ReportTheme.SECONDARY
    bc.bars.strokeWidth = 0.5
    bc.bars.strokeColor = colors.white
    
    bc.categoryAxis.categoryNames = [str(k) for k in data_dict.keys()]
    bc.categoryAxis.labels.angle = 45
    bc.categoryAxis.labels.boxAnchor = 'ne'
    bc.categoryAxis.labels.fontSize = 7
    bc.categoryAxis.labels.fontName = 'Helvetica'
    
    bc.valueAxis.valueMin = 0
    bc.valueAxis.labels.fontSize = 7
    bc.valueAxis.labels.fontName = 'Helvetica'
    
    drawing.add(bc)
    return drawing


def export_form_responses_to_pdf(form):
    buffer = io.BytesIO()
    # Usamos BaseDocTemplate para soportar encabezados y pies de página en cada página
    doc = BaseDocTemplate(buffer, pagesize=letter, 
                          leftMargin=50, rightMargin=50, topMargin=80, bottomMargin=60)
    
    # Definimos el área de contenido (Frame)
    frame = Frame(doc.leftMargin, doc.bottomMargin, doc.width, doc.height, id='normal')
    template = PageTemplate(id='main', frames=[frame], onPage=header_footer)
    doc.addPageTemplates([template])
    
    elements = []
    styles = ReportTheme.get_styles()

    # Estilos personalizados
    title_style = styles['ModernTitle']
    heading_style = styles['ModernHeading']
    subheading_style = styles['ModernSubHeading']
    normal_style = styles['ModernNormal']
    cell_style = styles['TableCell']

    # Recopilación de datos
    responses = (
        form.response_to
        .select_related('responder', 'responder__establecimiento')
        .prefetch_related('response__answer_to')
        .all()
    )
    questions = form.questions.all()
    total_responses = responses.count()

    # --- PORTADA / ENCABEZADO DE TÍTULO ---
    elements.append(Paragraph(f"Reporte de Resultados", title_style))
    elements.append(Paragraph(f"Formulario: {form.title}", subheading_style))
    
    # Resumen Ejecutivo Inicial
    summary_data = [
        ["Métrica", "Valor"],
        ["Fecha de Generación", datetime.now().strftime('%d/%m/%Y %H:%M')],
        ["Total de Respuestas", str(total_responses)],
        ["Estado del Formulario", "Público" if form.is_public else "Privado"]
    ]
    elements.append(Spacer(1, 10))
    elements.append(_create_styled_table(summary_data, col_widths=[2.5*inch, 2.5*inch]))
    elements.append(Spacer(1, 30))

    # Mapa de opciones por pregunta
    choices_by_question = {
        question.id: {str(choice.id): choice.choice for choice in question.choices.all()}
        for question in questions
    }

    # --- ESTADÍSTICAS GENERALES ---
    elements.append(Paragraph("1. Resumen General", heading_style))
    elements.append(Paragraph("Resultados agregados de toda la población encuestada.", normal_style))
    elements.append(Spacer(1, 12))

    for q in questions:
        if q.question_type in ['multiple choice', 'checkbox']:
            elements.append(Paragraph(f"Pregunta: {q.question}", subheading_style))
            
            counts = {}
            q_responses_count = 0
            for r in responses:
                ans = [a for a in r.response.all() if a.answer_to_id == q.id]
                if ans:
                    q_responses_count += 1
                for a in ans:
                    raw_val = a.answer
                    parts = raw_val.split(',') if q.question_type == 'checkbox' else [raw_val]
                    for p in parts:
                        p = p.strip()
                        if p:
                            choice_text = choices_by_question[q.id].get(p, p)
                            counts[choice_text] = counts.get(choice_text, 0) + 1

            if counts:
                # Tabla de datos
                table_data = [[Paragraph("Opción", cell_style), Paragraph("Recuento", cell_style), Paragraph("%", cell_style)]]
                for opt, count in counts.items():
                    pct = (count / q_responses_count * 100) if q_responses_count > 0 else 0
                    table_data.append([Paragraph(opt, cell_style), str(count), f"{pct:.1f}%"])
                
                elements.append(_create_styled_table(table_data, col_widths=[3.5*inch, 1*inch, 1*inch]))
                elements.append(Spacer(1, 15))

                # Gráfico
                try:
                    drawing = _create_pie_chart_drawing(counts) if len(counts) <= 6 else _create_bar_chart_drawing(counts)
                    if drawing:
                        elements.append(drawing)
                except Exception as e:
                    elements.append(Paragraph(f"Error en gráfico: {str(e)}", normal_style))
                
                elements.append(Spacer(1, 24))
            else:
                elements.append(Paragraph("Sin respuestas registradas.", normal_style))
                elements.append(Spacer(1, 12))
        else:
            elements.append(Paragraph(f"Pregunta: {q.question} (Respuesta Abierta)", subheading_style))
            ans_count = sum(1 for r in responses if any(a.answer_to_id == q.id for a in r.response.all()))
            elements.append(Paragraph(f"Total de respuestas recibidas: {ans_count}", normal_style))
            elements.append(Spacer(1, 12))

    # --- ESTADÍSTICAS POR ESTABLECIMIENTO ---
    elements.append(PageBreak())
    elements.append(Paragraph("2. Estadísticas por Establecimiento", heading_style))
    elements.append(Paragraph("Distribución de la participación y resultados por unidad organizacional.", normal_style))
    elements.append(Spacer(1, 12))

    resp_by_est = {}
    for r in responses:
        est = r.responder.establecimiento if r.responder and r.responder.establecimiento else None
        est_name = str(est) if est else "Sin Establecimiento"
        if est_name not in resp_by_est:
            resp_by_est[est_name] = []
        resp_by_est[est_name].append(r)

    # Tabla resumen de participación
    est_summary_data = [[Paragraph("Establecimiento", cell_style), Paragraph("Respuestas", cell_style), Paragraph("%", cell_style)]]
    for est_name, est_resps in sorted(resp_by_est.items()):
        pct = (len(est_resps) / total_responses * 100) if total_responses > 0 else 0
        est_summary_data.append([Paragraph(est_name, cell_style), str(len(est_resps)), f"{pct:.1f}%"])
    
    elements.append(_create_styled_table(est_summary_data, col_widths=[3.5*inch, 1*inch, 1*inch]))
    elements.append(Spacer(1, 15))

    # Gráfico de participación
    try:
        est_counts = {name: len(resps) for name, resps in sorted(resp_by_est.items())}
        drawing = _create_pie_chart_drawing(est_counts)
        if drawing:
            elements.append(drawing)
    except:
        pass

    elements.append(Spacer(1, 30))

    # Detalle por establecimiento
    elements.append(Paragraph("3. Desglose Detallado", heading_style))
    
    for est_name, est_resps in sorted(resp_by_est.items()):
        elements.append(Spacer(1, 10))
        elements.append(Paragraph(f"Unidad: {est_name}", subheading_style))
        elements.append(Paragraph(f"Total Participantes: {len(est_resps)}", normal_style))
        elements.append(Spacer(1, 10))

        for q in questions:
            if q.question_type in ['multiple choice', 'checkbox']:
                est_q_counts = {}
                est_q_resp_count = 0
                for r in est_resps:
                    ans = [a for a in r.response.all() if a.answer_to_id == q.id]
                    if ans:
                        est_q_resp_count += 1
                    for a in ans:
                        raw_val = a.answer
                        parts = raw_val.split(',') if q.question_type == 'checkbox' else [raw_val]
                        for p in parts:
                            p = p.strip()
                            if p:
                                choice_text = choices_by_question[q.id].get(p, p)
                                est_q_counts[choice_text] = est_q_counts.get(choice_text, 0) + 1
                
                if est_q_counts:
                    elements.append(Paragraph(f"Pregunta: {q.question}", normal_style))
                    est_table_data = [[Paragraph("Opción", cell_style), Paragraph("Cant.", cell_style), Paragraph("%", cell_style)]]
                    for opt, count in est_q_counts.items():
                        pct = (count / est_q_resp_count * 100) if est_q_resp_count > 0 else 0
                        est_table_data.append([Paragraph(opt, cell_style), str(count), f"{pct:.1f}%"])
                    
                    elements.append(_create_styled_table(est_table_data, col_widths=[3.5*inch, 0.8*inch, 0.8*inch]))
                    elements.append(Spacer(1, 10))

                    try:
                        drawing = _create_bar_chart_drawing(est_q_counts)
                        if drawing:
                            # Escalar un poco el gráfico para que no ocupe tanto en el detalle
                            drawing.width = 350
                            drawing.height = 180
                            elements.append(drawing)
                    except:
                        pass
                    
                    elements.append(Spacer(1, 15))
        
        elements.append(Spacer(1, 10))

    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()
    return pdf